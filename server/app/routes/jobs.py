from fastapi import APIRouter, Depends, HTTPException, Response, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from typing import List, Optional, Any, Dict
from datetime import datetime, date
import csv, io, os, zipfile, tempfile

from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import httpx
import datetime as dt
import pandas as pd

from app.deps import get_db
from app.schemas import CreateJob, JobOut, PhotoOut
from app.models import new_job
from app.services.storage_s3 import presign_url, get_bytes
from app.utils import normalize_phone, build_required_types_for_sector, type_label, sector_by_id

router = APIRouter()


def oid(obj):
    return str(obj["_id"]) if isinstance(obj.get("_id"), ObjectId) else obj.get("_id")


# ------------------------------------------------------------
# LIST JOBS
# ------------------------------------------------------------
def _job_to_out(doc: dict) -> JobOut:
    created = doc.get("createdAt")
    if isinstance(created, (datetime, date)):
        created = created.isoformat()

    return JobOut(
        id=str(doc["_id"]),
        workerPhone=doc["workerPhone"],
        siteId=doc["siteId"],
        sectors=[{
            "sector": int(s["sector"]),
            "requiredTypes": s.get("requiredTypes", []),
            "currentIndex": int(s.get("currentIndex", 0)),
            "status": s.get("status", "PENDING"),
        } for s in (doc.get("sectors") or [])],
        status=doc.get("status", "PENDING"),
        createdAt=created,
        macId=doc.get("macId"),
        rsnId=doc.get("rsnId"),
        azimuthDeg=doc.get("azimuthDeg"),
    )


@router.get("/jobs")
def list_jobs(db=Depends(get_db)) -> List[JobOut]:
    docs = list(db.jobs.find({}, sort=[("_id", -1)]))
    return [_job_to_out(d) for d in docs]


@router.get("/jobs/{job_id}")
def get_job(
    job_id: str,
    sector: Optional[int] = Query(None),
    db=Depends(get_db)
) -> Dict[str, Any]:
    """
    Return a single job + its photos.
    If ?sector=<n> is provided, only photos tagged with that sector are returned.
    All photos include a presigned HTTPS URL (s3Url) suitable for <img src="...">.
    """
    try:
        _id = ObjectId(job_id)
    except Exception:
        raise HTTPException(400, "Invalid job id")

    job = db.jobs.find_one({"_id": _id})
    if not job:
        raise HTTPException(404, "Job not found")

    # Build photo query
    photo_q: Dict[str, Any] = {"jobId": str(job["_id"])}
    if sector is not None:
        photo_q["sector"] = int(sector)

    # Oldest→newest so UI shows natural order
    photos = list(db.photos.find(photo_q).sort("_id", 1))

    out_photos = []
    for p in photos:
        docp: Dict[str, Any] = {
            "id": str(p["_id"]),
            "jobId": p.get("jobId"),
            "type": p.get("type"),
            "sector": p.get("sector"),
            "status": p.get("status"),
            "reason": p.get("reason") or [],
            "fields": p.get("fields") or {},
            "checks": p.get("checks") or {},
            "phash": p.get("phash"),
            "ocrText": p.get("ocrText"),
            "s3Key": p.get("s3Key"),
        }

        # presign the S3 key so the browser can load a private object
        key = p.get("s3Key")
        docp["s3Url"] = presign_url(key, expires=3600) if key else None

        out_photos.append(docp)

    return {
        "job": _job_to_out(job).model_dump(),
        "photos": out_photos,
    }


@router.post("/jobs", response_model=JobOut)
def create_or_extend_job(payload: CreateJob, db=Depends(get_db)) -> JobOut:
    """
    Create a job or add a new sector block to an existing (workerPhone + siteId) pair.
    """
    worker = payload.workerPhone.strip()
    site = payload.siteId.strip()
    sector = int(payload.sector)
    worker_phone = normalize_phone(worker)
    if not worker or not site:
        raise HTTPException(400, "workerPhone and siteId are required")

    base = db.jobs.find_one({"workerPhone": worker_phone, "siteId": site})

    sector_required = build_required_types_for_sector(sector)
    sector_block = {
        "sector": sector,
        "requiredTypes": sector_required,
        "currentIndex": 0,
        "status": "PENDING",
    }

    if not base:
        doc = {
            "workerPhone": worker_phone,
            "siteId": site,
            "sectors": [sector_block],
            "status": "PENDING",
            "createdAt": datetime.utcnow().isoformat(),
        }
        ins = db.jobs.insert_one(doc)
        doc["_id"] = ins.inserted_id
        return _job_to_out(doc)

    existing = sector_by_id(base.get("sectors", []), sector)
    if existing:
        raise HTTPException(409, f"Sector {sector} already exists for this worker/site")

    db.jobs.update_one({"_id": base["_id"]}, {"$push": {"sectors": sector_block}})
    updated = db.jobs.find_one({"_id": base["_id"]})
    return _job_to_out(updated)


# ------------------------------------------------------------
# PER-JOB CSV
# ------------------------------------------------------------
@router.get("/jobs/{job_id}/export.csv")
def export_csv(job_id: str, db=Depends(get_db)):
    try:
        job = db.jobs.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(404, "Invalid Job ID format")

    if not job:
        raise HTTPException(404, "Job not found")

    photos = list(db.photos.find({"jobId": job_id}))
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([
        "jobId", "workerPhone", "photoId", "type", "s3Key",
        "macId", "rsn", "azimuthDeg",
        "blurScore", "isDuplicate", "skewDeg", "hasLabelIds",
        "status", "reason"
    ])
    for p in photos:
        f = p.get("fields", {})
        c = p.get("checks", {})
        writer.writerow([
            job_id,
            job.get("workerPhone", ""),
            oid(p),
            p.get("type", ""),
            p.get("s3Key", ""),
            f.get("macId"),
            f.get("rsn"),
            f.get("azimuthDeg"),
            c.get("blurScore"),
            c.get("isDuplicate"),
            c.get("skewDeg"),
            c.get("hasLabelIds"),
            p.get("status"),
            "|".join(p.get("reason", [])),
        ])
    data = out.getvalue().encode("utf-8")
    headers = {"Content-Disposition": f'attachment; filename="job_{job_id}.csv"'}
    return Response(content=data, headers=headers, media_type="text/csv")


# ------------------------------------------------------------
# ALL JOBS CSV
# ------------------------------------------------------------
@router.get("/jobs/export.csv")
def export_jobs_csv(db=Depends(get_db)):
    out = io.StringIO()
    writer = csv.writer(out)

    headers = [
        "Job ID", "Worker", "Sectors", "Status",
        "MAC ID", "RSN ID", "Azimuth (deg)",
        "Created At", "Updated At",
    ]
    writer.writerow(headers)

    cur = db.jobs.find().sort("createdAt", -1)
    for job in cur:
        mac = job.get("macId") or ""
        rsn = job.get("rsnId") or ""
        az = job.get("azimuthDeg")
        az = f"{az:.1f}" if isinstance(az, (int, float)) else ""

        # Try to backfill MAC/RSN from latest LABELLING photo if missing
        if not mac or not rsn:
            lab = db.photos.find_one(
                {"jobId": str(job["_id"]), "type": "LABELLING"},
                sort=[("_id", -1)]
            )
            if lab:
                f = lab.get("fields") or {}
                mac = mac or f.get("macId", "")
                rsn = rsn or f.get("rsn", "")

        # Render sectors as comma list for this overview
        sectors_list = [str(s.get("sector")) for s in (job.get("sectors") or [])]
        sectors_str = ", ".join(sectors_list)

        row = [
            str(job["_id"]),
            job.get("workerPhone", ""),
            sectors_str,
            job.get("status", ""),
            mac, rsn, az,
            job.get("createdAt", ""),
            job.get("updatedAt", ""),
        ]
        writer.writerow(row)

    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="jobs.csv"'},
    )


# ------------------------------------------------------------
# XLSX (no images)
# ------------------------------------------------------------
@router.get("/jobs/{job_id}/export.xlsx")
def export_xlsx(job_id: str, db=Depends(get_db)):
    try:
        job = db.jobs.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(404, "Invalid Job ID format")
    if not job:
        raise HTTPException(404, "Job not found")

    photos = list(db.photos.find({"jobId": job_id}))

    wb = Workbook()
    ws = wb.active
    ws.title = "Photos"

    headers = [
        "jobId", "workerPhone", "photoId", "type", "s3Key",
        "macId", "rsn", "azimuthDeg",
        "blurScore", "isDuplicate", "skewDeg", "hasLabelIds",
        "status", "reason"
    ]
    ws.append(headers)

    for p in photos:
        f = p.get("fields", {})
        c = p.get("checks", {})
        ws.append([
            job_id,
            job.get("workerPhone", ""),
            oid(p),
            p.get("type", ""),
            p.get("s3Key", ""),
            f.get("macId", ""),
            f.get("rsn", ""),
            f.get("azimuthDeg", ""),
            c.get("blurScore", ""),
            c.get("isDuplicate", ""),
            c.get("skewDeg", ""),
            c.get("hasLabelIds", ""),
            p.get("status", ""),
            "|".join(p.get("reason", [])),
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="job_{job_id}.xlsx"'},
    )

@router.delete("/jobs/{job_id}", status_code=204)
def delete_job(job_id: str, db=Depends(get_db)):
    """Delete a job and all its photos from MongoDB (no file deletions)."""
    try:
        _id = ObjectId(job_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid job id")

    job = db.jobs.find_one({"_id": _id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Delete associated photos (if any)
    db.photos.delete_many({"jobId": {"$in": [job_id, _id]}})

    # Delete job itself
    db.jobs.delete_one({"_id": _id})

    return Response(status_code=204)


# ------------------------------------------------------------
# XLSX (with images)
# ------------------------------------------------------------
@router.get("/jobs/{job_id}/export_with_images.xlsx")
def export_xlsx_with_images(job_id: str, db=Depends(get_db)):
    try:
        job = db.jobs.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(404, "Invalid Job ID format")
    if not job:
        raise HTTPException(404, "Job not found")

    photos = list(db.photos.find({"jobId": {"$in": [job_id, ObjectId(job_id)]}}).sort("_id", 1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Photos"

    headers = [
        "jobId", "workerPhone", "photoId", "type", "macId", "rsn", "azimuthDeg",
        "blurScore", "isDuplicate", "skewDeg", "hasLabelIds", "status", "reason", "thumbnail"
    ]
    ws.append(headers)

    ws.column_dimensions["N"].width = 28  # thumbnail column
    thumb_max = 160

    def _clean_key(k: str | None) -> str | None:
        if not k:
            return None
        k = str(k)
        if k.startswith("s3://"):
            try:
                return k.split("/", 3)[-1].split("/", 1)[-1]
            except Exception:
                return None
        return k

    row_idx = 2
    for p in photos:
        f = p.get("fields", {}) or {}
        c = p.get("checks", {}) or {}

        ws.append([
            str(job["_id"]),
            job.get("workerPhone", ""),
            str(p.get("_id")),
            p.get("type", ""),
            f.get("macId", ""),
            f.get("rsn", ""),
            f.get("azimuthDeg", ""),
            c.get("blurScore", ""),
            c.get("isDuplicate", ""),
            c.get("skewDeg", ""),
            c.get("hasLabelIds", ""),
            p.get("status", ""),
            "|".join(p.get("reason", [])),
            "",  # thumbnail cell
        ])

        key = _clean_key(p.get("s3Key"))
        if not key:
            row_idx += 1
            continue

        try:
            url = presign_url(key, expires=3600)
            with httpx.Client(timeout=20) as client:
                r = client.get(url)
                r.raise_for_status()
                with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                    img = PILImage.open(io.BytesIO(r.content)).convert("RGB")
                    img.thumbnail((thumb_max, thumb_max))
                    img.save(tmp.name, "JPEG", quality=80)
                    xlimg = XLImage(tmp.name)
                    ws.add_image(xlimg, f"N{row_idx}")
        except Exception as ex:
            print(f"[XLSX] fetch failed for key={key}: {ex}")

        row_idx += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="job_{job_id}_with_images.xlsx"'},
    )



# ------------------------------------------------------------
# JOB ZIP (images)
# ------------------------------------------------------------
@router.get("/jobs/{job_id}/export.zip")
def export_job_zip(job_id: str, sector: int | None = Query(None, description="Optional sector filter"),
                   db=Depends(get_db)):
    # 1) Resolve job
    try:
        _id = ObjectId(job_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid job id")

    job = db.jobs.find_one({"_id": _id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # 2) Build photo query (allow string/ObjectId jobId for older docs)
    photo_q = {"jobId": {"$in": [job_id, _id]}}
    if sector is not None:
        photo_q["sector"] = int(sector)

    photos = list(db.photos.find(photo_q).sort("_id", 1))
    if not photos:
        raise HTTPException(status_code=404, detail="No photos for this job/sector")

    # Helper: strip accidental s3://bucket/ prefixes to get the raw key
    def _clean_key(k: str | None) -> str | None:
        if not k:
            return None
        k = str(k)
        if k.startswith("s3://"):
            try:
                # s3://bucket/key -> split once and take the key part
                return k.split("/", 3)[-1].split("/", 1)[-1]
            except Exception:
                return None
        return k

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in photos:
            p_sector = p.get("sector")
            folder = f"Sec{p_sector}" if p_sector is not None else "Unknown"

            base = (p.get("type") or "PHOTO").lower()
            key_raw = p.get("s3Key") or ""
            key = _clean_key(key_raw)

            # infer extension from key
            ext = ".jpg"
            if key:
                low = key.lower()
                for e in (".jpeg", ".jpg", ".png", ".webp"):
                    if low.endswith(e):
                        ext = e
                        break

            logical = f"{base}{ext}" if p_sector is None else f"sec{p_sector}_{base}{ext}"
            arcname = f"{folder}/{logical}"

            # Try localPath first (rare)
            lp = p.get("localPath")
            if lp and os.path.exists(lp):
                zf.write(lp, arcname=arcname)
                continue

            # Always presign fresh from key (do NOT trust stored s3Url)
            if key:
                try:
                    url = presign_url(key, expires=3600)
                    with httpx.Client(timeout=20) as client:
                        r = client.get(url)
                        r.raise_for_status()
                        zf.writestr(arcname, r.content)
                        continue
                except Exception as ex:
                    print(f"[ZIP] fetch failed for key={key}: {ex}")

            # fallback marker if missing/failed
            zf.writestr(arcname.replace(ext, "_MISSING.txt"), b"Missing or inaccessible image")

    mem.seek(0)
    fname = f'job_{job_id}_sec{sector}.zip' if sector is not None else f'job_{job_id}.zip'
    return StreamingResponse(
        mem,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )



# ------------------------------------------------------------
# TEMPLATE FOR SECTOR (single)
# ------------------------------------------------------------
@router.get("/jobs/templates/sector/{sector}")
def job_template(sector: int):
    types = build_required_types_for_sector(sector)
    return {
        "requiredTypes": types,
        "labels": {t: type_label(t) for t in types},
        "sector": sector,
    }


# ------------------------------------------------------------
# MANUAL EXPORT – SECTOR-WISE EXCEL (one sheet per sector)
# ------------------------------------------------------------
def _dt_or_none(s: str | None):
    if not s:
        return None
    # Accept YYYY-MM-DD
    return dt.datetime.strptime(s, "%Y-%m-%d")


@router.get("/exports/sector.xlsx")
def export_sector_xlsx(
    db=Depends(get_db),
    date_from: str | None = Query(None, description="YYYY-MM-DD"),
    date_to: str | None = Query(None, description="YYYY-MM-DD"),
):
    """
    Manual export – data grouped by sector, one sheet per sector.
    Works with multi-sector jobs; uses photo.sector.
    """
    q = {}
    if date_from or date_to:
        tmin = _dt_or_none(date_from) or dt.datetime.min
        tmax = _dt_or_none(date_to) or dt.datetime.max
        q["createdAt"] = {"$gte": tmin, "$lte": tmax}

    jobs = list(db.jobs.find(q))
    by_sector: Dict[int | None, list[dict]] = {}

    for j in jobs:
        job_id = str(j["_id"])
        worker = j.get("workerPhone")
        photos = list(db.photos.find({"jobId": job_id}))
        for p in photos:
            photo_sector = p.get("sector")
            f = p.get("fields", {}) or {}
            c = p.get("checks", {}) or {}
            base = (p.get("type") or "PHOTO").lower()
            # keep original extension if we can infer from key
            ext = ".jpg"
            key = p.get("s3Key", "")
            for e in (".jpeg", ".jpg", ".png", ".webp"):
                if key.lower().endswith(e):
                    ext = e
                    break
            logical = (f"sec{photo_sector}_{base}{ext}"
                       if photo_sector is not None else f"{base}{ext}")

            row = {
                "jobId": job_id,
                "workerPhone": worker,
                "sector": photo_sector,
                "photoId": str(p.get("_id")),
                "type": p.get("type"),
                "s3Key": key,
                "s3Url": presign_url(key) if key else None,
                "logicalName": logical,
                "macId": f.get("macId"),
                "rsn": f.get("rsn"),
                "azimuthDeg": f.get("azimuthDeg"),
                "blurScore": c.get("blurScore"),
                "isDuplicate": c.get("isDuplicate"),
                "skewDeg": c.get("skewDeg"),
                "status": p.get("status"),
                "reason": "|".join(p.get("reason") or []),
            }
            by_sector.setdefault(photo_sector, []).append(row)

    # Build workbook (one sheet per sector)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def write_sheet(title: str, rows: list[dict]):
        ws = wb.create_sheet(title=title)
        if not rows:
            ws.append(["No data"])
            return
        df = pd.DataFrame(rows)
        ws.append(list(df.columns))
        for _, r in df.iterrows():
            ws.append(list(r.values))

    # Known sectors first (sorted), then Unknown
    for s in sorted(k for k in by_sector.keys() if k is not None):
        write_sheet(f"Sec{s}", by_sector[s])
    if None in by_sector:
        write_sheet("Unknown", by_sector[None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        headers={"Content-Disposition": 'attachment; filename="export_sector.xlsx"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
